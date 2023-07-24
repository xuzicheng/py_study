#读取xls文件中的数据
import xlrd

file = "原表.xls"
wb = xlrd.open_workbook(file) #读取工作簿
ws = wb.sheets()[0] #选第一个工作表
data = []

for row in range(7, ws.nrows):
    name = ws.cell(row, 1).value.strip() #科室名称
    total1 = ws.cell(row, 2).value #总计
    total2 = ws.cell(row, 3).value #计
    avg = ws.cell(row, 20).value #平均每日人次
    
    info_list=[name,total1,total2,avg]

    if info_list[0] != "": #去除空数据
        data.append(info_list)

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment #设置单元格格式
thin = Side(border_style="thin", color="000000")#定义边框粗细及颜色

file = "统计表模板.xlsx"
wb = load_workbook(file)
ws = wb.active

#写入数据
for i in data:
    ws.append(i)
    
#设置字号，对齐，缩小字体填充，加边框
#Font(bold=True)可加粗字体
for row_number in range(3, ws.max_row+1):    
    ws.row_dimensions[row_number].height = 25 #设置行高
    for col_number in range(1,5):
        c = ws.cell(row=row_number,column=col_number)
        c.font = Font(size=11,bold=True)
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        c.alignment = Alignment(horizontal="center", vertical="center")
    
wb.save("统计表.xlsx")