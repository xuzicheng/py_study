import  os
work_path = os.getcwd() + "\\资料" #获取当前工作路径，指定“资料”文件夹
#获取路径下所有.xlsx文件，并存入列表
pathss=[] # 存储文件夹内所有文件的路径（包括子目录内的文件）
for root, dirs, files in os.walk(work_path): 
    path = [os.path.join(root, name) for name in files] #将目录和文件名连接起来，才是完整文件路径
    for i in range(len(path)): #遍历所有文件的地址
        if path[i].endswith(".xlsx"): #只提取后缀为xlsx的文件
            pathss.append(path[i])

#定义函数，获取Excel表格中的ID数据
from openpyxl import load_workbook #用于读取Excel中的信息
def Get_system_ID(file):
    wb = load_workbook(file)
    ws = wb.active

    ID_list = []
    for row in range(2,ws.max_row+1):
        ID = ws["A"+str(row)].value #ID信息在A列
        if ID != None: #过滤空值
            ID_list.append(ID)
    return ID_list

#获取数据，存入总列表
total_list = []
for file in pathss:
    info = Get_system_ID(file)
    total_list += info

#写入数据到新的excel表，并设置格式
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment #字体，颜色，对齐
from openpyxl.styles import PatternFill #单元格填充

wb = Workbook() #新建Excel工作簿
ws = wb.active #使用活动工作表
ws.column_dimensions['A'].width=18.5 #设定A列宽度
ws.cell(row=1,column=1,value="ID") #写入字段名
color_fill = PatternFill(fill_type='solid', fgColor="B3CFA1") #设置底色
ws.cell(row=1, column=1).fill = color_fill #填充底色

#批量从列表中提取数据并写入
for row in range(1,len(total_list)+1):
    ws.cell(row=row+1,column=1,value=total_list[row-1])

#设置字号及对齐
font_set = Font(name='Arial', size=9)
for i in range(1,ws.max_row+1):
    ws.cell(row=i,column=1).font = font_set
    ws.cell(row=i,column=1).alignment = Alignment(horizontal='left', vertical='center',shrink_to_fit = True)

wb.save(os.getcwd()+"\\ID.xlsx")

print(f"\n共获取到 {len(pathss)} 个 Excel表，共 {len(total_list)} 个ID。")