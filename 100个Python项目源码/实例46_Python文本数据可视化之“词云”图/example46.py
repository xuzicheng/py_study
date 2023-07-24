import numpy as np # numpy数据处理库
import wordcloud # 词云库
from PIL import Image # 图像处理库，用于读取背景图片
import matplotlib.pyplot as plt # 图像展示库，以便在notebook中显示图片
from openpyxl import load_workbook #读取词频Excel文件
import os #获取词频Excel文件路径

#将存好的Excel词频表读取成字典
path='词频'  #文件所在文件夹
files = [path+"\\"+i for i in os.listdir(path)] #获取文件夹下的文件名,并拼接完整路径
maskImage = np.array(Image.open('background.png')) # 定义词频背景图

for file in files:
    #将词频Excel文件读取为字典
    wb = load_workbook(file)
    ws = wb.active
    wordFreq = {}
    for i in range(2,ws.max_row+1):
        word = ws["A"+str(i)].value
        freq = ws["B"+str(i)].value
        wordFreq[word] = freq    
    
    #定义词云样式
    wc = wordcloud.WordCloud(
        font_path='C:/Windows/Fonts/simhei.ttf', # 设置字体
        mask= maskImage, # 设置背景图
        max_words=500, # 最多显示词数
        max_font_size=100) # 字号最大值
    
    #生成词云图
    wc.generate_from_frequencies(wordFreq) # 从字典生成词云
    #保存图片到指定文件夹
    wc.to_file("词云图\\{}.png".format(file.split("\\")[1][:4]))
    #在notebook中显示词云图
    plt.imshow(wc) # 显示词云
    plt.axis('off') # 关闭坐标轴
    plt.show() # 显示图像