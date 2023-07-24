#从Excel表中提取数据
from openpyxl import load_workbook
wb = load_workbook("收货数据.xlsx")
ws = wb.active
data=[]
for row in range(2,ws.max_row+1):
    seq = ws["A"+str(row)].value
    supplier = ws["B"+str(row)].value
    material_pn = ws["C"+str(row)].value
    material_model = ws["D"+str(row)].value
    desp = ws["E"+str(row)].value
    qty = ws["F"+str(row)].value
    date = ws["G"+str(row)].value.date()
    info = [seq, supplier, material_pn, material_model, desp, qty, date]
    data.append(info)

#定义数量加总函数
def Sum_list(list):
    s = 0
    for i in list: #累加列表中的所有数
        s+=i
    return s

#加总数量列的所有数字
qty_list=[]
for i in data:
    qty_list.append(i[5]) #数量在内层列表的第6个位置，索引是5
sum_qty= Sum_list(qty_list) #调用加总函数加总


#定义合并单元格的函数    
def Merge_cells(table,target_list,start_row,col):
    '''
    table: 是需要操作的表格
    target_list: 是目标列表，即含有重复数据的列表
    start_row: 是开始行，即表格中开始比对数据的行（需要将标题除开）
    col: 是需要处理数据的列
    '''
    start = 0 #开始行计数
    end = 0 #结束行计数
    reference = target_list[0] #设定基准，以列表中的第一个字符串开始
    for i in range(len(target_list)): #遍历列表
        if target_list[i] != reference: #开始比对，如果内容不同执行如下
            reference = target_list[i] #基准变成列表中下一个字符串
            end = i - 1 
            table.cell(start+start_row,col).merge(table.cell(end+start_row,col))
            start = end + 1
        if i == len(target_list) - 1: #遍历到最后一行，按如下操作
            end = i
            table.cell(start+start_row,col).merge(table.cell(end+start_row,col))

#数据提取即处理完毕后，就可以往Word的表格中写入数据了
from docx import Document
doc = Document("收货记录模板.docx")
#读取word文档中的第一个表格的第二和第三列除标题和尾部总数行的数据
table = doc.tables[0] #已确定是第一个表格，其索引是0
supplier = [] #存储供应商名称
pn = [] #存储物料编码
for i in data:
    supplier.append(i[1])
    pn.append(i[2])
#按需增加行，以便填写数据
for i in range(len(supplier)): #模板中已经有一行了，所以总共只需增加len(supplier)行
    table.add_row() 
#增加好行后先做合并单元格操作
Merge_cells(table,supplier,1,1) #开始合并行为2，索引为1；供应商名称是在2列，索引为1
Merge_cells(table,pn,1,2) #开始合并行为2，索引为1；物料编码是在3列，索引为2

#写入数据到表格
for row in range(1,len(supplier)+1):
    for col in range(7):        
        table.cell(row,col).text = str(data[row-1][col])

max_row = len(table.rows) #获取最大一行
qty_row = max_row-1 #确定需要写入加总数据的一行
table.cell(qty_row,5).merge(table.cell(qty_row,5)) #合并右下角用于填写数量的两个单元格
table.cell(qty_row,4).text = '总数：'
table.cell(qty_row,5).text = str(sum_qty)

doc.save("收货记录.docx")

