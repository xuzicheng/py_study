from openpyxl import load_workbook
wb = load_workbook("data\领料单（每日）.xlsx")
sheet_names=wb.get_sheet_names() #获得工作簿的所有工作表名
for sheet_name in sheet_names: #遍历每个工作表，更改A4单元格的数据
    ws=wb[sheet_name]
    ws['A4'].value="零件测试领料单" #直接将A4单元格的值改为需要的
wb.save("data\领料单（每日）-更改后.xlsx")
