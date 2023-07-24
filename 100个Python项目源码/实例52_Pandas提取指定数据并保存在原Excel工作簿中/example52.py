import pandas as pd
df = pd.read_excel("物料表.xlsx", header  = 2)

df500 = df[df["数量"]>500]

with pd.ExcelWriter('物料表.xlsx', mode = 'a' ,engine='openpyxl',
                    datetime_format='YYYY-MM-DD') as writer:
    df500.to_excel(writer, sheet_name='数量大于500',index = False)

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
thin = Side(border_style="thin", color="000000")#定义边框粗细及颜色

wb = load_workbook("物料表.xlsx")
ws = wb["数量大于500"]

# 调整列宽
ws.column_dimensions['A'].width = 12
ws.column_dimensions['C'].width = 15.5
ws.column_dimensions['G'].width = 10

#设置字号，对齐，缩小字体填充，加边框
for row_number in range(2, ws.max_row+1):
    for col_number in range(1,ws.max_column+1):
        c = ws.cell(row=row_number,column=col_number)
        c.font = Font(size=10)
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        c.alignment = Alignment(horizontal="left", vertical="center")
wb.save("物料表.xlsx")