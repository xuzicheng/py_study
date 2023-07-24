from openpyxl import load_workbook
wb = load_workbook("邮件地址.xlsx")
ws = wb.active
address = {}
for i in range(2,ws.max_row+1):
    dept = ws["A"+ str(i)].value
    to_add = ws["B"+ str(i)].value
    cc_add = ws["C"+ str(i)].value
    address[dept] = [to_add,cc_add]


import win32com.client as win32
def Send_mail(to_add, cc_add,file_path,dept):
    '''
    传入参数说明：
    to_add,收件人地址
    cc_add,抄送地址
    file_path,附件路径
    dept,部门名称
    '''
    outlook = win32.Dispatch('Outlook.Application') # 调用windows outlook应用
    mail = outlook.CreateItem(0) # 创建邮件
    mail.to = to_add #收件人
    mail.cc = cc_add #抄送人
    mail.Subject = "{}年假情况".format(dept) #主题
    mail.Attachments.Add(file_path) #添加附件。若有多个附件，则多调用几次即可
    mail.Body = '''Dear All,\n这是{}的年假情况，请查收！谢谢。\n\nBest regards!\n人事部 小李'''.format(dept)#正文内容
    mail.Send() #发送邮件

#发送邮件到各部门
for dept in address.keys():    
    to_add = address[dept][0]
    cc_add = address[dept][1]
    file_path = os.getcwd() + '\\年假_按部门\\年假情况_{}.xlsx'.format(dept)    
    Send_mail(to_add, cc_add,file_path,dept)  
    
print("邮件发送完成。")